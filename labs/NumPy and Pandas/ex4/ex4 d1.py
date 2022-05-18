import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import openpyxl

# 1 -------------------------
arr = []
with open('input (3).txt') as data:
    for i in data:
        #print(data)
        arr.append(i.split())
for i in range(len(arr)):
    summ_arr = 0
    for j in range(len(arr[i])):
        summ_arr += int(arr[i][j])
    #print(summ_arr, end = ' ')


# 2 -------------------------
arr = []
with open('input (2).txt') as data:
    arr = data.read()
arr = arr.split('\n')
arr.pop()
#print(arr, '\n')

new_arr = []
for i in range(len(arr)):
    string = arr[i].split(' ')
    if i%2==0:
        #print(string)
        for j in range(len(string)):
            new_str = string[j]
            if new_str[-1]==',' or new_str[-1]=='.':
                punct = new_str[-1]
                string[j] = new_str[-2::-1] + punct
            else:
                string[j] = new_str[::-1]
        string = ' '.join(string)
        #print(string)
        new_arr.append(string)
    else:
        string = ' '.join(string)
        new_arr.append(string)
new_arr = '\n'.join(new_arr)

#print('\n', new_arr, sep='')

#3 ------------------------
arr = []
with open('input (5).txt') as data:
    arr = data.read()
num_letter = 0
num_word = 0
check = True
for i in range(len(arr)):
    if arr[i].isalpha():
        num_letter += 1
    if arr[i].isalpha() and check:
        num_word += 1
        check = False
    elif arr[i].isalpha()==False:
        check = True

        
arr = arr.split('\n')
if arr[-1] == '':
    arr.pop(-1)
num_string = len(arr)
#print(num_letter, 'letters', num_word, 'words', num_string, 'lines')

#4 ---------------------------
data = pd.read_csv('input (1).csv', sep = ';')
data = data.describe()
min_data = data.loc['min']
#print(min_data)
min_price = min_data[min_data.idxmin()]
#print(min_data[data.loc['min']==min_price])


#5 ------------------------------
arr = []
'''for i in range(1, 1001):
    workbook = openpyxl.load_workbook(f"roga/{i}.xlsx")
    worksheet = workbook.active
    #arr.append([])
    human = worksheet.cell(row=2, column=2)
    zar = worksheet.cell(row=2, column=4)
    arr.append(human.value + ' ' + str(zar.value))
    #print(arr[i-1])
arr = sorted(arr)'''
#print(*arr, sep='\n')


#6 --------------------------------
data = pd.read_excel('salaries.xlsx', index_col=0)
data = pd.DataFrame(data)
data['Медиана'] = data.median(axis=1)
data.loc['Средняя зарплата'] = data.mean(axis=0)
profess = data.loc['Средняя зарплата'].idxmax()
region = data['Медиана'].idxmax()
print(data)
print(region, ', ', profess, sep='')














